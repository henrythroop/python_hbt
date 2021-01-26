#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Created on Tue Jan 26 11:07:37 2021

@author: hthroop
"""

import numpy as np
import xlrd as xlrd  # Reads .xls, not .xlsx

import matplotlib.pyplot as plt
import numpy as np
from fuzzywuzzy import fuzz
from fuzzywuzzy import process

import os.path
import pickle

import random
import csv
from scipy import stats   # For lin regress

from astropy.table import Table

from scipy.stats.stats import pearsonr

import hbt_short as hbt

import re
import glob
import sys

from openpyxl import load_workbook  # Reads .xlsx
from openpyxl import Workbook

# This program takes a directory full of Michael New's "High Risk / High Impact" spreadsheets.
# It then converts from various forms of voting (tally markers, individual votes, etc.) into one format.
# Then, it merges *all* of the files into one consolidated file.
# The output is an Excel file that can be easily searched, sorted, etc.
#
# HBT 26-Jan-2021

def tally2num(s):
    """

    Convert a list of votes and value into a single output value
    One could imagine different ways to do this -- e.g., mean, or half votes, etc.
    Here, I just do the simplest, cruest method: a median, rounded up.

    Parameters
    ----------
    votes : array of integers
        Array of all votes (e.g., [4,5,1])
    vals : array of strings
        Array of all possible vote values (e.g., ['High', 'Medium', 'Low'])

    Returns
    -------
    Median of votes, as a string

    Counts how many times 'X' is in the string, or returns the number.
    'XXX' → 4. 'X' → 1. '3' → 3.
    Return value is integer

    """
   
    if (s is None):
        return(0)

    if (s == ''):
        return(0)

    s = str(s).lower()

    if 'x' in s:
        out = s.count('x')
    else:
        out = int(s)
        
    return(out)    

def calc_median(votes, vals):
    """

    Convert a list of votes and value into a single output value
    One could imagine different ways to do this -- e.g., mean, or half votes, etc.
    Here, I just do the simplest, cruest method: a median, rounded up.

    Parameters
    ----------
    votes : array of integers
        Array of all votes (e.g., [4,5,1])
    vals : array of strings
        Array of all possible vote values (e.g., ['High', 'Medium', 'Low'])

    Returns
    -------
    Median of votes, as a string

    """

    
    index = np.argmax(votes)     # Take the index of the highest element. 
                                 # Then round toward zero, which in effet rounds score up.
    
    return(vals[index])


def process_risk_impact():
    
    # The main program. This loops over all the files in the directory, and then outputs one merged file

    # Set the directory name. Change this if needed.
    
    dir = '/Users/hthroop/Documents/HQ/SSW19/RiskImpactSurveys'
            
    arr_number = []
    arr_PI = []
    arr_name_panel = []
    arr_adj_impact = []
    arr_adj_risk = []

    # Get a list of all the Excel files in the directory
    
    file_out = 'SSW19_RiskImpact.xlsx'
        
    files = glob.glob(os.path.join(dir, '*.xlsx'))
    
    files_good = []
    
    # If a file has '~$' in it, then it's a temporary file. Ignore this. Also ignore the output file.

    for file in files:
        if (file_out not in file) and ('$' not in file):
            files_good.append(file)     
    
    files = files_good
        
    for file in files:
        print(f'Reading: {file}')
        name_panel = file.replace(dir, '').replace('RiskImpact Spreadsheet ', '').replace('.xlsx', '').replace('/','')
        
        wb = load_workbook(file)
        sheets = wb.sheetnames
        sheet = wb[sheets[0]]
        
        # Now read rows until we get to one that says 'Proposal #' in column A
        # We do this in caase someone has added extra rows or instructions above
        
        A = sheet['A']
        for i,cell in enumerate(range(len(A))):
            value = A[i].value  # NB: A[0] is row Cell A1

            if value is not None:
                if 'Proposal #' in value:
                    row_index_header = i+1  # When we grab this, we should call it row i+1
    
        # Now process the header row, and parse it to get the column headers
        # We do this in case someone has added extra columns
                    
        row_header = sheet[row_index_header]
        
        for i,cell in enumerate(range(len(row_header))):
            value = row_header[i].value
            if value is not None:
                if 'Proposal' in value:
                    column_index_number = i
    
                if 'PI' in value:
                    column_index_PI = i
    
                if 'High' in value:
                    column_index_high = i
    
                if 'Medium' in value:
                    column_index_medium = i
    
                if 'Low' in value:
                    column_index_low = i
    
                if 'Great' in value:
                    column_index_great = i
    
                if 'Some' in value:
                    column_index_some = i
    
                if 'Little' in value:
                    column_index_little = i
                    
        # Now that we have all this, process the file
    
        char_columns = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ' 
    
        for i in range(50):  # No more than this number of rows of valid data
            
            row = sheet[row_index_header+i+1] # Grab the row itself
            
            number = row[column_index_number].value
            PI     = row[column_index_PI].value  
    
            if PI is not None:   # Exclude empty rows. Empty cells are usually set to None, not ''.
    
                # print(f'High = {row[column_index_high].value}') 
                score_high  = tally2num(row[column_index_high].value)
                score_medium= tally2num(row[column_index_medium].value)
                score_low   = tally2num(row[column_index_low].value)
                
                score_great  = tally2num(row[column_index_great].value)
                score_some   = tally2num(row[column_index_some].value)
                score_little = tally2num(row[column_index_little].value)
    
            # Now calc the median of these. Take median, and round *up*
            
                votes_impact = [score_high, score_medium, score_low]
                votes_risk   = [score_great, score_some, score_little]
        
                adj_impact = calc_median(votes_impact,['High', 'Medium', 'Low'])
                adj_risk   = calc_median(votes_risk,['Great', 'Some', 'Little'])
                
                if (False):
                    print(f'{i}: {number} {PI} {score_high}/{score_medium}/{score_low}' + 
                          f'  {score_great}/{score_some}/{score_little}' + 
                          f'  {adj_impact}  {adj_risk} {name_panel}')
            
            # Finally, output these to the ouput arrays
                
                arr_number.append(number)
                arr_PI.append(PI)
                arr_name_panel.append(name_panel)
                arr_adj_impact.append(adj_impact)
                arr_adj_risk.append(adj_risk)
    
    # Now, put these all into an output file
    
    num_proposals = len(arr_number)
    
    wb = Workbook()
    ws = wb.active
    ws.title = "SSW19"
    
    # Make the header rows in the output file
        
    ws['A1'] = 'Proposal #'
    ws['B1'] = 'PI'
    ws['C1'] = 'Panel'
    ws['D1'] = 'Impact'
    ws['E1'] = 'Risk'
     
    # Fill out all the data in the output file
    
    for i in range(num_proposals):
        row = i + 2
        ws[f'A{row}'] = arr_number[i]
        ws[f'B{row}'] = arr_PI[i]
        ws[f'C{row}'] = arr_name_panel[i]
        ws[f'D{row}'] = arr_adj_impact[i]
        ws[f'E{row}'] = arr_adj_risk[i]

    # Write the ouput file
        
    file_out_full = os.path.join(dir, file_out)
    wb.save(filename = os.path.join(file_out_full))
    print(f'Wrote:   {file_out_full}')

# Wrapper routine to run the main function
        
if (__name__ == '__main__'):
    process_risk_impact()